""""""
import openpyxl

import logging
from logging.config import fileConfig

fileConfig('config\logging_config.ini')
logger = logging.getLogger(__name__)

class ExcelParser:
    """
    
    ---
    Attributes
    ----------


    Methods
    -------

    
    """
    name = "Excel Parser"
    version = "0.0.1"

    def __init__(self):
        pass

    def __str__(self):
        return "{} {}".format(self.name, self.version)
        
    def get_extensions(self):
        """
        
        Parameters
        ----------

        Returns
        -------
        None
        """
        return [".xlsx"]

    def get_functions(self):
        """
        
        Parameters
        ----------

        Returns
        -------
        None
        """
        return {
            "metadata": self.get_metadata,
            "analyse": self.analyze,
            "contents": self.get_contents
        }

    def get_metadata(self, filepath: str) -> dict:
        """This will return the doc info infomation from the 
        Named file.
        
        Parameters
        ----------

        Returns
        -------
        None
        """

        data = {}

        try:
            doc = openpyxl.load_workbook(filepath)

            # get the core properties from the file...
            # https://python-docx.readthedocs.io/en/latest/api/document.html#coreproperties-objects
            cp = doc.properties

            data['author'] = cp.creator
            data['category'] = cp.category
            data['comments'] = cp.description
            data['created'] = cp.created
            data['identifier'] = cp.identifier
            data['keywords'] = cp.keywords
            data['language'] = cp.language
            data['last_modified_by'] = cp.lastModifiedBy
            data['last_printed'] = cp.lastPrinted
            data['modified'] = cp.modified
            data['revision'] = cp.revision
            data['subject'] = cp.subject
            data['title'] = cp.title
            data['version'] = cp.version
            
        except Exception as ex:
             logger.error("{}\n\t>>>{}".format(filepath,ex))
        
        return data

    def analyze(self, filepath: str) -> dict:
        """
        
        Parameters
        ----------

        Returns
        -------
        None
        """
        data = dict
        
        return data

    def get_contents(self, filepath: str) -> list:
        """This will return the paragroah objects in a word document
        
        Parameters
        ----------

        Returns
        -------
        None
        """
        data = []
        try:
            wb = openpyxl.load_workbook(filepath)
            for sname in wb.sheetnames:
                ws = wb[sname]
                for values in ws.iter_rows(min_row=ws.min_row, 
                              max_row=ws.max_row, 
                              min_col=ws.min_column, 
                              max_col=ws.max_column,
                              values_only=True):
                    data.append("\t".join(map(str, values)))
            
        except Exception as ex:
            print("ERROR: {}\n\t>>>{}".format(filepath,ex))

        return {
            "type": "rows",   
            "data": data
        }